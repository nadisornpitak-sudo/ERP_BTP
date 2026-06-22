"""
Import All Stock sheets from 'สรุป STOCK  17.06.2569.xlsx' and Reorganize Racks in BTP ERP
"""
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

# Paths
SERVER_DIR = Path(__file__).parent
DATA_DIR = SERVER_DIR / "data"
DB_FILE = DATA_DIR / "btp_erp.json"
ROOT_DB_FILE = SERVER_DIR.parent / "data" / "erp_db.json"
XLSX_FILE = Path("C:/Users/User/Desktop/สรุป STOCK  17.06.2569.xlsx")

# Racks definition matching the frontend
WH3D_RACKS = [
    # Floor 1
    {'id':'A1','bays':1,'levels':4,'pos':5},
    {'id':'A2','bays':1,'levels':4,'pos':5},
    {'id':'A3','bays':1,'levels':4,'pos':5},
    {'id':'B1','bays':1,'levels':4,'pos':5},
    {'id':'B7','bays':1,'levels':4,'pos':5},
    {'id':'B8','bays':1,'levels':4,'pos':5},
    {'id':'B2','bays':1,'levels':4,'pos':5},
    {'id':'B3','bays':1,'levels':4,'pos':5},
    {'id':'B4','bays':1,'levels':4,'pos':5},
    {'id':'B5','bays':1,'levels':4,'pos':5},
    {'id':'B6','bays':1,'levels':4,'pos':5},
    {'id':'Y1','bays':1,'levels':4,'pos':5},
    {'id':'Y2','bays':1,'levels':4,'pos':5},
    {'id':'W1','bays':5,'levels':4,'pos':3},
    # Floor 2 Room 1
    {'id':'P1','bays':1,'levels':4,'pos':5},
    {'id':'P2','bays':1,'levels':4,'pos':5},
    {'id':'P3','bays':1,'levels':4,'pos':5},
    {'id':'P4','bays':2,'levels':4,'pos':5},
    {'id':'P5','bays':1,'levels':4,'pos':5},
    {'id':'P6','bays':1,'levels':4,'pos':5},
    {'id':'P7','bays':2,'levels':4,'pos':5},
    {'id':'P8','bays':2,'levels':4,'pos':5},
    # Floor 2 Room 2
    {'id':'Q1','bays':1,'levels':4,'pos':5},
    {'id':'Q2','bays':1,'levels':4,'pos':5},
    {'id':'Q3','bays':1,'levels':4,'pos':5},
    {'id':'Q4','bays':2,'levels':4,'pos':5},
    {'id':'Q5','bays':2,'levels':4,'pos':5},
    {'id':'Q6','bays':1,'levels':4,'pos':5},
    {'id':'Q7','bays':1,'levels':4,'pos':5},
    {'id':'Q8','bays':1,'levels':4,'pos':5},
    # Extension Floor 2 Room 1
    {'id':'X1','bays':1,'levels':4,'pos':5},
    {'id':'X2','bays':1,'levels':4,'pos':5},
    {'id':'X3','bays':1,'levels':4,'pos':5},
    {'id':'X4','bays':1,'levels':4,'pos':5},
    {'id':'X5','bays':1,'levels':4,'pos':5},
]

def generate_slots():
    slots = []
    for r in WH3D_RACKS:
        bays = r['bays']
        levels = r['levels']
        pos = r['pos']
        for l in range(levels):
            for b in range(bays):
                for p in range(pos):
                    slots.append(f"{r['id']}-{b+1}-{l+1}-{p+1}")
    return slots

def to_int(val):
    if not val or str(val).strip() in ['', '-', ' -']:
        return 0
    try:
        return int(float(str(val).replace(',', '').replace(' ', '').strip()))
    except ValueError:
        return 0

def main():
    if not XLSX_FILE.exists():
        print(f"❌ Error: ไม่พบไฟล์ {XLSX_FILE}")
        sys.exit(1)
        
    if not DB_FILE.exists():
        print(f"❌ Error: ไม่พบไฟล์ฐานข้อมูล BTP ERP ที่ {DB_FILE}")
        sys.exit(1)
        
    # Generate slots
    all_slots = generate_slots()
    print(f"📦 Generate Slots คลังสินค้าสำเร็จ: {len(all_slots)} พิกัดชั้นวาง")
    
    # 1. Parse Excel file (All sheets)
    try:
        import openpyxl
    except ImportError:
        print("❌ Error: กรุณาติดตั้ง openpyxl ก่อน: pip install openpyxl")
        sys.exit(1)
        
    print(f"📂 กำลังเปิดไฟล์: {XLSX_FILE.name}")
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    
    excel_data = {} # sku -> {s, d, o, total}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"📄 กำลังอ่าน Sheet: {sheet_name}")
        count = 0
        for r in range(6, sheet.max_row + 1):
            sku = sheet.cell(r, 2).value
            if not sku:
                continue
            sku = str(sku).strip()
            if not sku.startswith("FG"):
                continue
                
            s_val = to_int(sheet.cell(r, 4).value)
            d_val = to_int(sheet.cell(r, 5).value)
            o_val = to_int(sheet.cell(r, 6).value)
            rem_val = to_int(sheet.cell(r, 7).value)
            
            # Fallback to sum if remaining is 0 or not set
            if rem_val == 0:
                rem_val = s_val + d_val + o_val
                
            if rem_val > 0:
                excel_data[sku] = {
                    "S": s_val,
                    "D": d_val,
                    "O": o_val,
                    "total": rem_val
                }
                count += 1
        print(f"   -> พบรายการสต็อก {count} รายการ")
        
    print(f"\n📊 รวมข้อมูลจาก Excel สำเร็จ: พบยอดสต็อก {len(excel_data)} SKUs รวมทั้งสิ้น {sum(v['total'] for v in excel_data.values()):,} ชิ้น")
    
    # 2. Load current ERP database
    with open(DB_FILE, "r", encoding="utf-8") as f:
        db = json.load(f)
        
    db.setdefault("locations", {})
    prod_map = {p["sku"]: p for p in db.get("products", [])}
    
    # Clean invalid locations from locations map
    slot_set = set(all_slots)
    to_delete = []
    for sku, slot in db["locations"].items():
        if "floor" in slot.lower():
            continue
        if slot not in slot_set:
            to_delete.append(sku)
    for sku in to_delete:
        del db["locations"][sku]
        
    # 3. Update stock quantities from Excel into database products dict
    updated_stock_count = 0
    sku_not_found = []
    
    for sku, info in excel_data.items():
        if sku not in prod_map:
            sku_not_found.append(sku)
            continue
            
        prod = prod_map[sku]
        prod["stock"] = {
            "S": info["S"],
            "D": info["D"],
            "O": info["O"],
            "W": 0,
            "Q": 0,
            "F": 0
        }
        updated_stock_count += 1
        
    print(f"📈 อัปเดตยอดสต็อกลงสินค้าใน ERP สำเร็จ: {updated_stock_count} SKUs")
    
    # 4. Check existing locations and keep ONLY those that match the floor rules for their dominant channel
    occupied = set()
    valid_locations = {}
    mismatched_locations_count = 0
    
    # Sort products to ensure consistent slot assignment
    products_sorted = sorted(db.get("products", []), key=lambda x: (x.get("group", ""), x.get("sku", "")))
    
    for p in products_sorted:
        sku = p["sku"]
        stock_dict = p.get("stock", {})
        total_stock = sum(int(v) for v in stock_dict.values())
        
        if total_stock <= 0:
            # Clear location for items with no stock to free up racks
            if sku in db["locations"]:
                mismatched_locations_count += 1
            continue
            
        # Get dominant channel
        info = excel_data.get(sku)
        dominant = "O"
        if info:
            max_val = max(info["S"], info["D"], info["O"])
            if max_val > 0:
                if max_val == info["D"]:
                    dominant = "D"
                elif max_val == info["S"]:
                    dominant = "S"
                    
        # Check current slot
        current_slot = db["locations"].get(sku)
        if current_slot:
            if "floor" in current_slot.lower():
                # Keep custom floor locations if it's Floor 1 / Online
                if dominant == "O":
                    valid_locations[sku] = current_slot
                else:
                    mismatched_locations_count += 1
            elif dominant == "D" and current_slot.startswith("Q"):
                # Dealer item correctly on Q rack
                valid_locations[sku] = current_slot
                occupied.add(current_slot)
            elif dominant == "S" and current_slot.startswith(("P", "X")):
                # Shop item correctly on P/X rack
                valid_locations[sku] = current_slot
                occupied.add(current_slot)
            elif dominant == "O" and current_slot.startswith(("A", "B", "Y", "W")):
                # Online item correctly on Floor 1 rack
                valid_locations[sku] = current_slot
                occupied.add(current_slot)
            else:
                # Wrong floor or wrong channel, clear it so it gets re-allocated to the correct floor
                mismatched_locations_count += 1
                
    db["locations"] = valid_locations
    print(f"🧹 เคลียร์ตำแหน่งเดิมที่อยู่ผิดโซน/ผิดชั้นสำเร็จ: {mismatched_locations_count} รายการ")
    
    # 5. Allocate correct rack slots to items without locations
    assigned_count = 0
    fallback_assigned_count = 0
    
    for p in products_sorted:
        sku = p["sku"]
        stock_dict = p.get("stock", {})
        total_stock = sum(int(v) for v in stock_dict.values())
        
        if total_stock > 0 and sku not in db["locations"]:
            # Get dominant channel
            info = excel_data.get(sku)
            dominant = "O"
            if info:
                max_val = max(info["S"], info["D"], info["O"])
                if max_val > 0:
                    if max_val == info["D"]:
                        dominant = "D"
                    elif max_val == info["S"]:
                        dominant = "S"
                        
            # Determine preferred zone prefixes
            if dominant == "D":
                preferred_prefixes = ("Q",)
            elif dominant == "S":
                preferred_prefixes = ("P", "X")
            else:
                preferred_prefixes = ("A", "B", "Y", "W")
                
            # Try to assign to preferred zone
            free_slot = None
            for s in all_slots:
                if s not in occupied and s.startswith(preferred_prefixes):
                    free_slot = s
                    break
                    
            # Fallback to any free slot in the warehouse
            if not free_slot:
                for s in all_slots:
                    if s not in occupied:
                        free_slot = s
                        break
                if free_slot:
                    fallback_assigned_count += 1
            else:
                assigned_count += 1
                
            if free_slot:
                db["locations"][sku] = free_slot
                occupied.add(free_slot)
                
    print(f"📍 จัดตำแหน่งแร็คในโซนที่เหมาะสมสำเร็จ: {assigned_count} SKU")
    if fallback_assigned_count > 0:
        print(f"   ⚠️ ต้องจัดใส่โซนสำรองเนื่องจากโซนหลักเต็ม: {fallback_assigned_count} SKU")
        
    # 6. Distribute stock to channels based on assigned slot location (strict layout enforcement)
    floor1_count = 0
    floor2_r2_count = 0
    floor2_r1_count = 0
    
    for p in db.get("products", []):
        sku = p["sku"]
        stock_dict = p.get("stock", {})
        total_stock = sum(int(v) for v in stock_dict.values())
        
        if total_stock <= 0:
            continue
            
        slot = db["locations"].get(sku)
        if not slot:
            # Fallback to Online (O) if no slot found
            p["stock"] = {"S": 0, "D": 0, "O": total_stock, "W": 0, "Q": 0, "F": 0}
            floor1_count += 1
            continue
            
        # Re-distribute stock channels strictly
        if slot.startswith(("A", "B", "Y", "W")) or "floor" in slot.lower():
            # Floor 1 (Online)
            p["stock"] = {"S": 0, "D": 0, "O": total_stock, "W": 0, "Q": 0, "F": 0}
            floor1_count += 1
        elif slot.startswith("Q"):
            # Floor 2 Room 2 (Dealer)
            p["stock"] = {"S": 0, "D": total_stock, "O": 0, "W": 0, "Q": 0, "F": 0}
            floor2_r2_count += 1
        elif slot.startswith(("P", "X")):
            # Floor 2 Room 1 / Ext (Shop/หน้าร้าน)
            p["stock"] = {"S": total_stock, "D": 0, "O": 0, "W": 0, "Q": 0, "F": 0}
            floor2_r1_count += 1
        else:
            # Fallback
            p["stock"] = {"S": 0, "D": 0, "O": total_stock, "W": 0, "Q": 0, "F": 0}
            floor1_count += 1

    print("\n📈 ผลการจัดแจงช่องทางสต็อกตามตำแหน่ง:")
    print(f"   • ชั้น 1 (Online - O)          : {floor1_count} SKUs")
    print(f"   • ชั้น 2 ห้อง 2 (Dealer - D)     : {floor2_r2_count} SKUs")
    print(f"   • ชั้น 2 ห้อง 1 (หน้าร้าน - S)    : {floor2_r1_count} SKUs")
    
    # Calculate channel stocks inside DB now
    total_S = sum(p.get("stock", {}).get("S", 0) for p in db.get("products", []))
    total_D = sum(p.get("stock", {}).get("D", 0) for p in db.get("products", []))
    total_O = sum(p.get("stock", {}).get("O", 0) for p in db.get("products", []))
    print(f"\n📊 ยอดสต็อกหลังจัดแจงช่องทาง:")
    print(f"   • Shop (S): {total_S:,} ชิ้น")
    print(f"   • Dealer (D): {total_D:,} ชิ้น")
    print(f"   • Online (O): {total_O:,} ชิ้น")
    print(f"   • รวมทั้งหมด: {total_S + total_D + total_O:,} ชิ้น")
    
    # Save DB with backup
    backup_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = DATA_DIR / f"backup_pre_import_all_{backup_ts}.json"
    shutil.copy2(DB_FILE, backup_file)
    print(f"\n💾 สำรองข้อมูลเดิมไว้ที่: {backup_file.name}")
    
    # Write to local server DB
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, separators=(",", ":"))
    print("✅ บันทึกข้อมูลลงฐานข้อมูลเซิร์ฟเวอร์ (BTP-ERP-Server/data/btp_erp.json) สำเร็จ!")
    
    # Write to root workspace DB
    ROOT_DB_FILE.parent.mkdir(exist_ok=True)
    with open(ROOT_DB_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, separators=(",", ":"))
    print("✅ บันทึกข้อมูลลงฐานข้อมูลโฟลเดอร์หลัก (data/erp_db.json) สำเร็จ!")
    
    print("\n💡 สามารถรัน python upload_data_to_railway.py เพื่อนำข้อมูลที่จัดแจงใหม่นี้ขึ้นระบบคลาวด์ได้ทันที")

if __name__ == "__main__":
    main()
