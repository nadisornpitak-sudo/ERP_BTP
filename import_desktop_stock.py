"""
Import ONLY "คงเหลือ" column from Desktop Stock_Dashboard.xlsx to BTP ERP
"""
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

# Paths
DESKTOP_DIR = Path("c:/Users/User/Desktop")
XLSX_FILE = DESKTOP_DIR / "Stock_Dashboard.xlsx"
SERVER_DIR = Path(__file__).parent
DATA_DIR = SERVER_DIR / "data"
DB_FILE = DATA_DIR / "btp_erp.json"

def main():
    if not XLSX_FILE.exists():
        print(f"❌ Error: ไม่พบไฟล์ Stock_Dashboard.xlsx บนหน้า Desktop ที่ {XLSX_FILE}")
        sys.exit(1)
        
    if not DB_FILE.exists():
        print(f"❌ Error: ไม่พบไฟล์ฐานข้อมูล BTP ERP ที่ {DB_FILE}")
        sys.exit(1)
        
    print(f"📂 กำลังเปิดไฟล์: {XLSX_FILE}")
    try:
        import openpyxl
    except ImportError:
        print("❌ Error: กรุณาติดตั้ง openpyxl ก่อน:")
        print("   pip install openpyxl")
        sys.exit(1)
        
    # Open workbook with data_only=True to evaluate formulas (like =H4+I4-J4)
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    
    sheet_name = "SKU_Detail"
    if sheet_name not in wb.sheetnames:
        # Fallback to whatever sheet name matches similar text or use first sheet
        matched = [s for s in wb.sheetnames if "detail" in s.lower() or "sku" in s.lower()]
        if matched:
            sheet_name = matched[0]
        else:
            sheet_name = wb.sheetnames[0]
            
    print(f"📄 ใช้ข้อมูลจาก Sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    
    # Iterate and find headers
    col_sku_idx = -1
    col_rem_idx = -1
    
    # We look in the first 5 rows to find headers
    header_row_num = -1
    for r_idx in range(1, 6):
        row_vals = [cell.value for cell in ws[r_idx]]
        # Find SKU
        for i, val in enumerate(row_vals):
            if val and str(val).strip() in ["รหัสสินค้า", "SKU", "รหัส"]:
                col_sku_idx = i
                header_row_num = r_idx
                break
        if col_sku_idx != -1:
            # Find Remaining
            for i, val in enumerate(row_vals):
                if val and str(val).strip() in ["คงเหลือ", "Remaining", "สต็อกคงเหลือ", "คงเหลือสุทธิ"]:
                    col_rem_idx = i
                    break
            break
            
    if col_sku_idx == -1 or col_rem_idx == -1:
        print(f"❌ Error: ไม่พบหัวตาราง รหัสสินค้า (SKU) หรือ คงเหลือ (Remaining)")
        print(f"   ตรวจพบคีย์ค้นหา SKU ในแถวที่ {header_row_num} แต่ไม่พบคอลัมน์คงเหลือ")
        sys.exit(1)
        
    print(f"🔍 หัวตารางในแถวที่ {header_row_num}:")
    print(f"   • รหัสสินค้า (SKU) คอลัมน์ที่: {col_sku_idx + 1}")
    print(f"   • คงเหลือ คอลัมน์ที่: {col_rem_idx + 1}")
    
    # Read rows starting from header_row_num + 1
    excel_stocks = {}
    for r_idx in range(header_row_num + 1, ws.max_row + 1):
        sku = ws.cell(r_idx, col_sku_idx + 1).value
        rem_val = ws.cell(r_idx, col_rem_idx + 1).value
        
        if sku is None:
            continue
            
        sku_str = str(sku).strip()
        if not sku_str or sku_str == "nan" or sku_str.startswith("#"):
            continue
            
        # Parse stock count
        try:
            if rem_val is None or rem_val == "":
                qty = 0
            else:
                qty = int(float(rem_val))
        except (ValueError, TypeError):
            qty = 0
            
        excel_stocks[sku_str] = qty
        
    print(f"📊 ตรวจพบข้อมูลใน Excel ทั้งหมด {len(excel_stocks)} รายการ")
    
    # Load current ERP database
    with open(DB_FILE, "r", encoding="utf-8") as f:
        db = json.load(f)
        
    prod_map = {p["sku"]: p for p in db.get("products", [])}
    
    updated_count = 0
    sku_not_found = []
    
    for sku, qty in excel_stocks.items():
        if sku not in prod_map:
            sku_not_found.append(sku)
            continue
            
        prod = prod_map[sku]
        # Set stock directly to Online (O) and clear others to match exact "คงเหลือ" total
        prod["stock"] = {
            "S": 0,
            "D": 0,
            "O": qty,  # Put entire quantity in Online (O)
            "W": 0,
            "Q": 0,
            "F": 0
        }
        updated_count += 1
        
    print(f"📈 ผลการซิงค์ข้อมูลสต็อก:")
    print(f"   • อัปเดตยอดสต็อกสำเร็จ: {updated_count} SKUs")
    if sku_not_found:
        print(f"   ⚠️ ไม่พบรหัสสินค้าในระบบ ERP ({len(sku_not_found)} SKUs): {', '.join(sku_not_found[:10])}...")
        
    # Save database with backup
    backup_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = DATA_DIR / f"backup_pre_desktop_import_{backup_ts}.json"
    shutil.copy2(DB_FILE, backup_file)
    print(f"💾 สำรองข้อมูลเดิมไว้ที่: {backup_file.name}")
    
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, separators=(",", ":"))
        
    print("✅ ซิงค์ยอดสต็อกคงเหลือเข้าระบบเรียบร้อยแล้ว!")
    print("\n💡 สามารถรัน python upload_data_to_railway.py เพื่อนำยอดสต็อกนี้ขึ้นระบบคลาวด์ได้ทันที")

if __name__ == "__main__":
    main()
